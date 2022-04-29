# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import json
from datetime import datetime, timedelta
from math import isclose
from urllib import request

from openpyxl import load_workbook


def check_nbp_price(date):
    prev_date = date - timedelta(days=1)
    day_string = prev_date.strftime("%Y-%m-%d")

    url = 'http://api.nbp.pl/api/exchangerates/rates/a/usd/%s/' % (day_string)

    try:
        with request.urlopen(url) as response:
            body = json.loads(response.read())
            return body['rates'][0]['mid']
    except:
        new_date = prev_date - timedelta(days=1)
        return check_nbp_price(new_date)


def parse_saxo_trades():
    transactions = []

    file = 'modified_apha_tilray_merge_saxo_report.xlsx'
    # file = 'saxo_trades_executed.xlsx'
    wb = load_workbook(filename=file)

    trades = wb['Trades']

    itertrades = iter(trades.rows)
    next(itertrades)
    for row in itertrades:
        if row[0].value == 'SWAP':
            transactions.append((row[0].value, row[1].value, row[2].value, row[3].value))
        else:
            date = row[3].value
            instrument = row[2].value
            amount = row[6].value
            booked_amount = row[10].value

            transactions.append((instrument, date, amount, booked_amount, booked_amount / amount))

    return transactions


def parse_revolut_trades():
    transactions = []

    deposit = []
    withdrawal = []
    dividend = []
    fee = []
    split = []

    wb = load_workbook(filename='revolut_trades_executed.xlsx')

    trades = wb['in']

    itertrades = iter(trades.rows)
    next(itertrades)
    for row in itertrades:
        split_row = row[0].value.split(',')
        (date, instrument, operation, amount, per_share, total, currency, fx_rate) = split_row

        if operation == 'STOCK SPLIT':
            split.append(row)
            transactions.append(("SPLIT", instrument, date, float(amount)))
        elif operation == 'CASH TOP-UP':
            deposit.append((date, float(total)))
        elif operation == 'CUSTODY_FEE':
            fee.append((date, float(total)))
        elif operation == 'DIVIDEND':
            dividend.append((date, instrument, float(total)))
        elif operation == 'CASH WITHDRAWAL':
            withdrawal.append((date, float(total)))
        else:
            if operation == 'SELL':
                amount = -float(amount)
            transactions.append((instrument, datetime.strptime(date, "%d/%m/%Y %H:%M:%S"), float(amount), float(total),
                                 float(total) / float(amount)))

    dep_sum = 0
    for dep in deposit:
        dep_sum += dep[1]

    print("Revolut deposited total of: %.2f" % dep_sum)

    wit_sum = 0
    for wit in withdrawal:
        wit_sum += abs(wit[1])

    print("Revolut withdrew total of %.2f" % wit_sum)

    return transactions


def process_transactions(transactions):
    to_date = {}
    to_tax = {}
    profit = {}
    cost = {}

    for transaction in transactions:
        if transaction[0] == "SPLIT":
            print("TESLA SPLIT")
            continue

        if transaction[0] == "SWAP":
            print("SWAP EVENT")

            (method, from_instrument, to_instrument, remaining_amount) = transaction

            swap_from_existing_positions = to_date[from_instrument]

            swap_booked = 0
            swap_amount = 0
            for position in swap_from_existing_positions[1]:
                swap_amount += position[0]
                swap_booked += position[0] * position[1]

            stack = [(remaining_amount, swap_booked / remaining_amount, swap_booked)]

            del to_date[from_instrument]
            to_date[to_instrument] = (remaining_amount, stack)
            continue

        (instrument, date, amount, booked_amount, per_share) = transaction

        if date.year not in to_tax:
            to_tax[date.year] = []
            profit[date.year] = []
            cost[date.year] = []

        if instrument not in to_date:
            to_date[instrument] = (0, [])

        (amount_owned, price_stack) = to_date[instrument]

        if amount > 0:
            amount_owned += amount
            price_stack.append((amount, per_share, booked_amount))

            to_date[instrument] = (amount_owned, price_stack)
        else:
            print("Selling %d of %s" % (abs(amount), instrument))
            print(to_date[instrument])
            sold_amount = amount
            tax = 0
            while not isclose(amount, 0, abs_tol=0.000001):
                if len(price_stack) == 0:
                    print(
                        "------------------- Taxation event with 0 shares. Check for errors --------------------------")
                    break

                (owned_amount, bought_at_price_per_share, buy_booked_cost) = price_stack.pop(0)

                bought_at = bought_at_price_per_share
                sold_at = per_share
                per_share_taxable = abs(sold_at) - abs(bought_at)

                if abs(owned_amount) <= abs(amount):
                    amount += abs(owned_amount)

                    if amount > 0:
                        price_stack.insert(0, (amount, bought_at_price_per_share, amount * bought_at_price_per_share))

                    tax += per_share_taxable * abs(owned_amount)
                    cost[date.year].append((instrument, abs(bought_at_price_per_share) * abs(owned_amount)))
                    profit[date.year].append((instrument, abs(sold_at) * abs(owned_amount)))

                else:
                    new_trans_amount = abs(owned_amount) - abs(amount)
                    amount = 0
                    new_trans_booked_amount = new_trans_amount * bought_at_price_per_share

                    amount_sold = abs(owned_amount) - abs(new_trans_amount)
                    tax += per_share_taxable * amount_sold
                    cost[date.year].append((instrument, abs(bought_at_price_per_share) * abs(amount_sold)))
                    profit[date.year].append((instrument, abs(sold_at) * abs(amount_sold)))

                    price_stack.insert(0, (new_trans_amount, bought_at_price_per_share, new_trans_booked_amount))

            to_tax[date.year].append((sold_amount, instrument, per_share, tax))

            print("Taxation event sold %d of %s at %.2f per share. Taxable income is %.2f" % (
            abs(sold_amount), instrument, abs(per_share), tax))

    return (to_tax, profit, cost)


def apply_exchange_rate(transaction):
    if transaction[0] != "SPLIT" and transaction[0] != 'SWAP':
        (instrument, date, amount, booked_amount, per_share) = transaction
        exchange_rate = check_nbp_price(date)

        return instrument, date, amount, booked_amount * exchange_rate, per_share * exchange_rate
    else:
        return transaction


if __name__ == '__main__':

    revolut_transaction_list = parse_revolut_trades()
    saxo_transaction_list = parse_saxo_trades()

    transaction_list = revolut_transaction_list + saxo_transaction_list
    transaction_list = list(map(apply_exchange_rate, transaction_list))

    (taxable_events, profit, cost) = process_transactions(transaction_list)

    for year in taxable_events:
        print("Calculating year %d" % year)
        final_year_taxable_income = 0
        cost_year = 0
        profit_year = 0

        for cost_event in cost[year]:
            cost_year += cost_event[1]

        for profit_event in profit[year]:
            profit_year += profit_event[1]

        for event in taxable_events[year]:
            print("Event %s" % str(event))
            final_year_taxable_income += event[3]

        final_year_tax = final_year_taxable_income * 0.19

        print("%d :Final cost %.2f" % (year, cost_year))
        print("%d :Final profit %.2f" % (year, profit_year))

        print("%d :Final taxable income %.2f" % (year, final_year_taxable_income))
        print("%d :Final tax %.2f" % (year, final_year_tax))
